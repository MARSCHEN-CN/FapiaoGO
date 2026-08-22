"""
Excel / CSV 导出模块（性能优化版）

优化点：
- openpyxl 使用 write_only=True 模式，大数量导出快 5-10 倍
- 样式对象复用，不逐单元格重复创建
- 数据行构建使用列表推导和预计算，减少函数调用开销
- 正则预编译（已在模块级完成）
"""

import re
import os
import csv
import logging
from pathlib import Path
from decimal import Decimal
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.worksheet.cell_range import CellRange
except ImportError:
    Workbook = Font = PatternFill = Border = Side = Alignment = None
    get_column_letter = WriteOnlyCell = CellRange = None

logger = logging.getLogger(__name__)

_PATTERN_BUYER_SELLER = re.compile(r'\[(?:BUYER|SELLER)_(?:START|END)\]')
_PATTERN_AUX = re.compile(r'__AUX_[A-Za-z0-9_]+__')
_PATTERN_WHITESPACE = re.compile(r'\s+')

EXCEL_FORMULA_PREFIXES = ('=', '+', '-', '@')
EXPORT_ALLOWED_EXTENSIONS = {'xlsx', 'csv'}

EXPORT_ALLOWED_BASE_DIRS = [
    str(Path.home()),
    str(Path.home() / 'Desktop'),
    str(Path.home() / 'Documents'),
    str(Path.home() / 'Downloads'),
]

ALLOWED_EXPORT_KEYS = {
    'serialNo', 'invoiceType', 'invoiceDate', 'invoiceNumber',
    'amountWithoutTax', 'taxAmount', 'totalAmount',
    'buyerName', 'buyerTaxNo', 'sellerName', 'sellerTaxNo', 'issuer',
    'classificationCode', 'xmmc', 'ggxh', 'unit', 'quantity',
    'unitPrice', 'lineAmount', 'taxRate', 'lineTax',
    'note', 'originalFilename',
}

MASTER_WIDTH = {
    'serialNo': 8, 'invoiceType': 12, 'invoiceDate': 12, 'invoiceNumber': 20,
    'amountWithoutTax': 12, 'taxAmount': 12, 'totalAmount': 12,
    'buyerName': 25, 'buyerTaxNo': 20, 'sellerName': 25, 'sellerTaxNo': 20, 'issuer': 10,
    'classificationCode': 18, 'xmmc': 35, 'ggxh': 20, 'unit': 8,
    'quantity': 10, 'unitPrice': 12, 'lineAmount': 12, 'taxRate': 12,
    'lineTax': 12, 'note': 30, 'originalFilename': 35,
}

INVOICE_LEVEL_KEYS = {
    'serialNo', 'invoiceType', 'invoiceDate', 'invoiceNumber',
    'amountWithoutTax', 'taxAmount', 'totalAmount',
    'buyerName', 'buyerTaxNo', 'sellerName', 'sellerTaxNo', 'issuer',
    'note', 'originalFilename',
}

MONEY_KEYS = {'amountWithoutTax', 'taxAmount', 'totalAmount', 'unitPrice', 'lineAmount', 'lineTax'}

_COLS_CACHE = {}

def _get_default_cols(include_remark):
    """获取默认列定义（缓存）。"""
    key = ('default', include_remark)
    if key not in _COLS_CACHE:
        cols = [
            ('序号', 'serialNo', 8, False),
            ('发票类型', 'invoiceType', 12, False),
            ('开票日期', 'invoiceDate', 12, False),
            ('发票号码', 'invoiceNumber', 20, False),
            ('税前金额', 'amountWithoutTax', 12, False),
            ('税额合计', 'taxAmount', 12, False),
            ('价税合计', 'totalAmount', 12, False),
            ('购买方名称', 'buyerName', 25, False),
            ('购买方税号', 'buyerTaxNo', 20, False),
            ('销售方名称', 'sellerName', 25, False),
            ('销售方税号', 'sellerTaxNo', 20, False),
            ('开票人', 'issuer', 10, False),
            ('分类编码', 'classificationCode', 18, False),
            ('项目名称', 'xmmc', 35, False),
            ('规格型号', 'ggxh', 20, False),
            ('单位', 'unit', 8, False),
            ('数量', 'quantity', 10, False),
            ('单价', 'unitPrice', 12, False),
            ('金额', 'lineAmount', 12, False),
            ('税率/征收率', 'taxRate', 12, False),
            ('税额', 'lineTax', 12, False),
        ]
        if include_remark:
            cols.extend([('备注', 'note', 30, False), ('原文件名', 'originalFilename', 35, False)])
        _COLS_CACHE[key] = cols
    return _COLS_CACHE[key]

def _get_cols_from_columns(columns):
    """将客户端列定义转为内部元组格式（缓存）。"""
    if not columns:
        return None
    key = tuple(c['key'] for c in columns)
    if key not in _COLS_CACHE:
        _COLS_CACHE[key] = [(c['label'], c['key'], int(c.get('width') or MASTER_WIDTH.get(c['key'], 12)), bool(c.get('virtual'))) for c in columns]
    return _COLS_CACHE[key]

def _invoice_identity(rec):
    """Invoice Entity Boundary Freeze v1: invoiceDocumentId 为领域主键，最高优先级。
    
    invoiceDocumentId > recordId > originalFilename > invoiceNumber > __ANON
    invoiceNumber 仅作为最终兜底，不作为身份主键。
    """
    return (
        rec.get('invoiceDocumentId') or
        rec.get('recordId') or
        rec.get('originalFilename') or
        rec.get('invoiceNumber') or
        f"__ANON_{id(rec)}"
    )


def sanitize_columns(columns):
    if not columns:
        return None
    out = []
    for c in columns:
        if not isinstance(c, dict):
            continue
        key = c.get('key')
        if key not in ALLOWED_EXPORT_KEYS:
            continue
        out.append({
            'key': key,
            'label': c.get('label') or key,
            'width': int(c.get('width') or MASTER_WIDTH.get(key, 12)),
            'virtual': bool(c.get('virtual')),
        })
    return out or None


def _is_path_traversal(path: str) -> bool:
    if not path:
        return False
    if '..' in path:
        if path.startswith('..'):
            return True
        if path.startswith('/..') or path.startswith('\\..'):
            return True
        if '/../' in path or '\\..\\' in path:
            return True
        if path.endswith('/..') or path.endswith('\\..'):
            return True
    normalized = os.path.normpath(path)
    return '..' in normalized.split(os.sep)


def _is_path_inside_bases(path: str, bases: list) -> bool:
    try:
        abs_path = os.path.abspath(os.path.normpath(path))
        for base in bases:
            abs_base = os.path.abspath(os.path.normpath(base))
            if not abs_base.endswith(os.sep):
                abs_base += os.sep
            if abs_path.startswith(abs_base):
                return True
        return False
    except (OSError, ValueError):
        return False


def validate_export_path(file_path, fmt):
    if not isinstance(file_path, str) or not file_path.strip():
        raise ValueError("缺少 filePath 参数")
    raw_path = os.path.expanduser(file_path.strip())
    if _is_path_traversal(raw_path):
        raise ValueError("检测到路径遍历攻击")
    if not os.path.isabs(raw_path):
        raise ValueError("导出路径必须是绝对路径")
    normalized_path = os.path.abspath(raw_path)
    if not _is_path_inside_bases(normalized_path, EXPORT_ALLOWED_BASE_DIRS):
        raise ValueError(f"导出路径必须在以下目录内: {', '.join(EXPORT_ALLOWED_BASE_DIRS)}")
    requested_ext = str(fmt).lower().lstrip('.')
    actual_ext = os.path.splitext(normalized_path)[1].lower().lstrip('.')
    if requested_ext not in EXPORT_ALLOWED_EXTENSIONS:
        raise ValueError("不支持的导出格式")
    if actual_ext != requested_ext:
        raise ValueError(f"导出文件扩展名必须为 .{requested_ext}")
    parent_dir = os.path.dirname(normalized_path)
    if not parent_dir or not os.path.isdir(parent_dir):
        raise ValueError("导出目录不存在")
    return normalized_path, requested_ext


def sanitize_excel_text(value):
    if not isinstance(value, str):
        return value
    value = _PATTERN_BUYER_SELLER.sub(' ', value)
    value = _PATTERN_AUX.sub(' ', value)
    value = _PATTERN_WHITESPACE.sub(' ', value).strip()
    if value and value[0] in EXCEL_FORMULA_PREFIXES:
        return f"'{value}"
    return value


def sanitize_export_row(row):
    return [sanitize_excel_text(v) for v in row]


def sanitize_sheet_name(name, max_len=31):
    sanitized = re.sub(r'[\[\]\*\/\\\?:]', '_', name)[:max_len]
    return sanitized or 'Sheet'


def _safe_num(val):
    if val is None or val == '':
        return 0.0
    try:
        return float(str(val).replace(',', ''))
    except (ValueError, TypeError):
        return 0.0


def _quantity_value(raw):
    """数量字段：整数保持字符串（防 "001" → 1），小数转float。"""
    if raw is None or raw == '':
        return ''
    q_str = str(raw)
    try:
        q_num = float(q_str.replace(',', ''))
        return q_str if q_num == int(q_num) else q_num
    except (ValueError, TypeError):
        return sanitize_excel_text(raw)


class _XlsxWriteOnlyWriter:
    """write_only 模式 XLSX 写入器（性能优化版）。"""

    def __init__(self, include_remark, on_progress):
        self.include_remark = include_remark
        self.on_progress = on_progress
        self.money_fmt = '#,##0.00'

        thin_side = Side(style='thin')
        self.thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        self.header_align = Alignment(vertical='center', horizontal='center')
        self.data_align = Alignment(vertical='center')
        self.right_align = Alignment(vertical='center', horizontal='right')
        self.total_font = Font(bold=True, size=11)
        self.total_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        self.total_border = Border(top=Side(style='thin'), bottom=Side(style='double'))

    def _make_cell(self, ws, value, *, font=None, fill=None, alignment=None, border=None, number_format=None):
        cell = WriteOnlyCell(ws, value=value)
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border
        if number_format: cell.number_format = number_format
        return cell

    def _build_data_row(self, ws, inv, cols, serial):
        row = []
        for _, key, _, virtual in cols:
            if key == 'serialNo' or virtual:
                val = serial
                cell = self._make_cell(ws, val, border=self.thin_border, alignment=self.data_align)
            elif key in MONEY_KEYS:
                num = _safe_num(inv.get(key, ''))
                if isinstance(num, (int, float)):
                    cell = self._make_cell(ws, num, border=self.thin_border, alignment=self.right_align, number_format=self.money_fmt)
                else:
                    cell = self._make_cell(ws, sanitize_excel_text(inv.get(key, '')), border=self.thin_border, alignment=self.data_align)
            elif key == 'quantity':
                cell = self._make_cell(ws, _quantity_value(inv.get(key, '')), border=self.thin_border, alignment=self.data_align)
            else:
                cell = self._make_cell(ws, sanitize_excel_text(inv.get(key, '')), border=self.thin_border, alignment=self.data_align)
            row.append(cell)
        return row

    def write_sheet(self, ws, sheet_invoices, sheet_label='', columns=None):
        on_progress = self.on_progress
        sheet_total = len(sheet_invoices)
        if sheet_total == 0:
            ws.append([])
            return

        if columns:
            cols = _get_cols_from_columns(columns)
            group_key_fn = _invoice_identity
        else:
            cols = _get_default_cols(self.include_remark)
            group_key_fn = _invoice_identity  # Step 5B: 统一使用 _invoice_identity，不再裸用 invoiceNumber

        col_index_map = {key: c for c, (_, key, _, _) in enumerate(cols, 1)}

        for c, (header, _, width, _) in enumerate(cols, 1):
            cell = self._make_cell(ws, header, font=self.header_font, fill=self.header_fill,
                                   alignment=self.header_align, border=self.thin_border)
            ws.column_dimensions[get_column_letter(c)].width = width
            if c == 1:
                header_row = [cell]
            else:
                header_row.append(cell)
        ws.append(header_row)

        group_map = defaultdict(list)
        for inv in sheet_invoices:
            group_map[group_key_fn(inv)].append(inv)

        update_interval = max(1, sheet_total // 15)
        serial = 1
        row_idx = 2
        written = 0
        merge_ranges = []

        for group in group_map.values():
            start_row = row_idx
            row_count = len(group)
            for inv in group:
                ws.append(self._build_data_row(ws, inv, cols, serial))
                row_idx += 1
                written += 1
                if on_progress and written % update_interval == 0:
                    on_progress(15 + int(55 * written / sheet_total), 100,
                               f'写入{sheet_label} ({written}/{sheet_total})...')
            serial += 1
            if row_count > 1:
                end_row = start_row + row_count - 1
                for ikey in INVOICE_LEVEL_KEYS:
                    ci = col_index_map.get(ikey)
                    if ci:
                        merge_ranges.append(f'{get_column_letter(ci)}{start_row}:{get_column_letter(ci)}{end_row}')

        seen_invoices = set()
        sums = {k: Decimal('0') for k in ('amountWithoutTax', 'taxAmount', 'totalAmount', 'lineAmount', 'lineTax')}
        for group in group_map.values():
            for inv in group:
                inv_id = group_key_fn(inv)
                if inv_id not in seen_invoices:
                    seen_invoices.add(inv_id)
                    for k in ('amountWithoutTax', 'taxAmount', 'totalAmount'):
                        v = _safe_num(inv.get(k, 0))
                        if v is not None:
                            sums[k] += Decimal(str(v))
                for k in ('lineAmount', 'lineTax'):
                    v = _safe_num(inv.get(k, 0))
                    if v is not None:
                        sums[k] += Decimal(str(v))

        # 合计值按「列 key」匹配（与前端 computeTotals 语义一致）。
        # 不依赖中文列标签：前端自定义列的 label 可能是「总税额/总金额」
        # （见 frontend/src/export/excelColumns.js），而默认列用「税额合计/价税合计」，
        # 按 label 匹配会导致合计行对应单元格留空（回归：总税额/总金额未合计）。
        total_value_map = {
            'amountWithoutTax': sums['amountWithoutTax'],
            'taxAmount': sums['taxAmount'],
            'totalAmount': sums['totalAmount'],
            'lineAmount': sums['lineAmount'],
            'lineTax': sums['lineTax'],
        }
        total_row = []
        for c, (_, key, _, _) in enumerate(cols, 1):
            if c == 1:
                cell = self._make_cell(ws, '合计', font=self.total_font, fill=self.total_fill, border=self.total_border)
            elif key in total_value_map:
                v = total_value_map[key]
                if isinstance(v, Decimal):
                    v = float(round(v, 2))
                    cell = self._make_cell(ws, v, font=self.total_font, fill=self.total_fill,
                                           border=self.total_border, alignment=self.right_align, number_format=self.money_fmt)
                else:
                    cell = self._make_cell(ws, v, font=self.total_font, fill=self.total_fill, border=self.total_border)
            else:
                cell = self._make_cell(ws, None, fill=self.total_fill, border=self.total_border)
            total_row.append(cell)
        ws.append(total_row)

        for mr in merge_ranges:
            ws.merged_cells.ranges.add(CellRange(mr))

        ws.freeze_panes = 'A2'
        last_col = get_column_letter(len(cols))
        ws.auto_filter.ref = f'A1:{last_col}1'


def export_csv(file_path, invoices, options, on_progress=None):
    include_remark = options.get('includeRemark', True)
    columns = options.get('columns')
    total = len(invoices)

    if columns:
        cols = _get_cols_from_columns(columns)
        headers = [c[0] for c in cols]
        ordered = list(invoices)
        group_key_fn = _invoice_identity
    else:
        default_cols = _get_default_cols(include_remark)
        headers = [c[0] for c in default_cols]
        cols = default_cols
        ordered = sorted(invoices, key=lambda x: x.get('invoiceNumber', ''))
        group_key_fn = lambda inv: inv.get('invoiceNumber', '')

    if on_progress:
        on_progress(10, 100, f'正在写入 CSV ({total} 行)...')
    update_interval = max(1, total // 20)

    with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(headers)

        groups = defaultdict(list)
        for inv in ordered:
            groups[group_key_fn(inv)].append(inv)

        serial = 1
        written = 0
        for group in groups.values():
            for inv in group:
                row = []
                for _, key, _, virtual in cols:
                    if key == 'serialNo' or virtual:
                        row.append(serial)
                    else:
                        row.append(inv.get(key, ''))
                writer.writerow(sanitize_export_row(row))
                written += 1
                if on_progress and written % update_interval == 0:
                    on_progress(10 + int(85 * written / max(total, 1)), 100, f'写入 CSV ({written}/{total})...')
            serial += 1

    if on_progress:
        on_progress(95, 100, 'CSV 写入完成')


def export_xlsx(file_path, invoices, options, on_progress=None):
    if Workbook is None:
        raise RuntimeError("需要安装 openpyxl: pip install openpyxl")

    include_remark = options.get('includeRemark', True)
    split_by_type = options.get('splitByType', False)
    columns = options.get('columns')

    if on_progress:
        on_progress(5, 100, '创建 Excel 工作簿...')

    wb = Workbook(write_only=True)

    writer = _XlsxWriteOnlyWriter(include_remark=include_remark, on_progress=on_progress)

    if split_by_type:
        type_groups = defaultdict(list)
        for inv in invoices:
            type_groups[inv.get('invoiceType', '未知类型')].append(inv)

        if on_progress:
            on_progress(10, 100, f'分类整理 ({len(type_groups)} 种类型)...')

        used_names = set()
        for type_name, type_invs in type_groups.items():
            name = sanitize_sheet_name(type_name, 27)
            if name in used_names:
                i = 2
                while f'{name}({i})' in used_names:
                    i += 1
                name = f'{name}({i})'
            used_names.add(name)
            ws = wb.create_sheet(title=name)
            writer.write_sheet(ws, type_invs, sheet_label=f'[{name}]', columns=columns)
    else:
        ws = wb.create_sheet(title='发票汇总')
        writer.write_sheet(ws, invoices, sheet_label='发票汇总', columns=columns)

    if on_progress:
        on_progress(95, 100, '正在保存文件...')

    wb.save(file_path)

    if on_progress:
        on_progress(100, 100, '导出完成')
