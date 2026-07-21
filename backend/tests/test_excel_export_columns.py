"""Excel 导出「列驱动」后端契约测试（Commit 1）。

覆盖 v2.1 验收目标：
1. /api/export-excel-rows 同源 rows（路由在 app.py，本测试覆盖 exporter 内核）
2. columns=[invoiceNumber,buyerName] → 仅两列、顺序即传入顺序
3. serialNo 为 virtual，不触发 KeyError
4. 合计逻辑：invoice 级去重 + line 级全累加
5. 未传 columns 时，旧导出行为不变
"""
import os
import sys
import csv
import tempfile
import unittest

# 将 backend/ 加入路径，便于直接运行本文件
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import excel_exporter as ex  # noqa: E402

try:
    from openpyxl import load_workbook
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False


def _sample_invoices():
    """模拟 _db_record_to_export 输出的扁平行（2 张发票 3 行明细）。"""
    return [
        # 发票 A：2 行明细
        {'invoiceNumber': 'INV-001', 'file_name': 'a.pdf', 'buyerName': 'BuyerA',
         'amountWithoutTax': 300, 'taxAmount': 30, 'totalAmount': 330,
         'xmmc': '商品A1', 'lineAmount': 100, 'lineTax': 10},
        {'invoiceNumber': 'INV-001', 'file_name': 'a.pdf', 'buyerName': 'BuyerA',
         'amountWithoutTax': 300, 'taxAmount': 30, 'totalAmount': 330,
         'xmmc': '商品A2', 'lineAmount': 200, 'lineTax': 20},
        # 发票 B：1 行明细
        {'invoiceNumber': 'INV-002', 'file_name': 'b.pdf', 'buyerName': 'BuyerB',
         'amountWithoutTax': 50, 'taxAmount': 5, 'totalAmount': 55,
         'xmmc': '商品B1', 'lineAmount': 50, 'lineTax': 5},
    ]


class TestSanitizeColumns(unittest.TestCase):
    def test_none_and_empty(self):
        self.assertIsNone(ex.sanitize_columns(None))
        self.assertIsNone(ex.sanitize_columns([]))
        self.assertIsNone(ex.sanitize_columns([{}]))

    def test_whitelist_drops_unknown_and_preserves_order(self):
        cols = [
            {'key': 'invoiceNumber'},
            {'key': 'password'},          # 越权 key，应丢弃
            {'key': 'buyerName', 'virtual': False},
        ]
        out = ex.sanitize_columns(cols)
        self.assertEqual([c['key'] for c in out], ['invoiceNumber', 'buyerName'])

    def test_preserves_incoming_order_no_sort(self):
        # 故意乱序传入，验证不排序（尊重用户字段选择顺序）
        cols = [
            {'key': 'buyerName'},
            {'key': 'invoiceNumber'},
            {'key': 'lineAmount'},
        ]
        out = ex.sanitize_columns(cols)
        self.assertEqual([c['key'] for c in out], ['buyerName', 'invoiceNumber', 'lineAmount'])

    def test_virtual_passthrough_and_width_fallback(self):
        cols = [{'key': 'serialNo', 'virtual': True}]
        out = ex.sanitize_columns(cols)
        self.assertTrue(out[0]['virtual'])
        self.assertEqual(out[0]['width'], ex.MASTER_WIDTH['serialNo'])

    def test_issuer_whitelisted(self):
        # 开票人 (issuer) 必须被白名单放行，否则勾选后预览有、导出没有
        cols = ex.sanitize_columns([{'key': 'issuer', 'label': '开票人'}])
        self.assertEqual([c['key'] for c in cols], ['issuer'])
        self.assertEqual(cols[0]['width'], ex.MASTER_WIDTH['issuer'])


class TestInvoiceIdentity(unittest.TestCase):
    def test_invoice_number_first(self):
        self.assertEqual(ex._invoice_identity({'invoiceNumber': 'X'}), 'X')

    def test_file_name_fallback(self):
        self.assertEqual(ex._invoice_identity({'file_name': 'f.pdf'}), 'f.pdf')

    def test_anon_fallback_stable(self):
        rec = {}
        a = ex._invoice_identity(rec)
        b = ex._invoice_identity(rec)
        self.assertTrue(a.startswith('__ANON_'))
        self.assertEqual(a, b)  # 同一对象 → 稳定（非每次新 Symbol）


class TestExportCsvColumns(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix='.csv', delete=False)
        self.tmp.close()
        self.path = self.tmp.name

    def tearDown(self):
        os.unlink(self.path)

    def test_only_selected_columns_in_order(self):
        cols = ex.sanitize_columns([
            {'key': 'buyerName', 'label': '购买方名称'},
            {'key': 'invoiceNumber', 'label': '发票号码'},
        ])
        ex.export_csv(self.path, _sample_invoices(),
                      {'includeRemark': False, 'columns': cols})
        with open(self.path, encoding='utf-8-sig') as f:
            rows = list(csv.reader(f))
        self.assertEqual(rows[0], ['购买方名称', '发票号码'])
        # 3 数据行
        self.assertEqual(len(rows), 4)

    def test_serial_virtual_no_keyerror(self):
        cols = ex.sanitize_columns([
            {'key': 'serialNo', 'virtual': True, 'label': '序号'},
            {'key': 'invoiceNumber', 'label': '发票号码'},
            {'key': 'lineAmount', 'label': '金额'},
        ])
        ex.export_csv(self.path, _sample_invoices(),
                      {'includeRemark': False, 'columns': cols})
        with open(self.path, encoding='utf-8-sig') as f:
            rows = list(csv.reader(f))
        # 表头
        self.assertEqual(rows[0], ['序号', '发票号码', '金额'])
        # 发票 A 两行共享序号 1；发票 B 序号 2
        self.assertEqual(rows[1][0], '1')
        self.assertEqual(rows[2][0], '1')
        self.assertEqual(rows[3][0], '2')
        # 金额按行输出
        self.assertEqual(rows[1][2], '100')
        self.assertEqual(rows[2][2], '200')
        self.assertEqual(rows[3][2], '50')

    def test_unknown_key_dropped_at_export(self):
        cols = ex.sanitize_columns([
            {'key': 'invoiceNumber', 'label': '发票号码'},
            {'key': 'not_a_real_field'},
        ])
        # sanitize 已丢弃越权 key，导出只含发票号码
        self.assertEqual([c['key'] for c in cols], ['invoiceNumber'])
        ex.export_csv(self.path, _sample_invoices(),
                      {'includeRemark': False, 'columns': cols})
        with open(self.path, encoding='utf-8-sig') as f:
            rows = list(csv.reader(f))
        self.assertEqual(rows[0], ['发票号码'])

    def test_issuer_exports_in_csv(self):
        # 开票人列必须真正出现在导出（验证白名单放行后链路打通）
        cols = ex.sanitize_columns([
            {'key': 'invoiceNumber', 'label': '发票号码'},
            {'key': 'issuer', 'label': '开票人'},
        ])
        ex.export_csv(self.path, _sample_invoices(),
                      {'includeRemark': False, 'columns': cols})
        with open(self.path, encoding='utf-8-sig') as f:
            rows = list(csv.reader(f))
        self.assertEqual(rows[0], ['发票号码', '开票人'])

    def test_backward_compat_no_columns(self):
        # 不传 columns → 旧默认列（含备注/原文件名当 includeRemark=True）
        ex.export_csv(self.path, _sample_invoices(), {'includeRemark': True})
        with open(self.path, encoding='utf-8-sig') as f:
            rows = list(csv.reader(f))
        header = rows[0]
        self.assertIn('发票号码', header)
        self.assertIn('购买方名称', header)
        self.assertIn('备注', header)
        self.assertIn('原文件名', header)


@unittest.skipIf(not HAVE_OPENPYXL, 'openpyxl 未安装，跳过 XLSX 断言')
class TestExportXlsxColumns(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        self.tmp.close()
        self.path = self.tmp.name

    def tearDown(self):
        os.unlink(self.path)

    def test_columns_order_and_virtual_serial(self):
        cols = ex.sanitize_columns([
            {'key': 'invoiceNumber', 'label': '发票号码'},
            {'key': 'buyerName', 'label': '购买方名称'},
            {'key': 'lineAmount', 'label': '金额'},
            {'key': 'amountWithoutTax', 'label': '税前金额'},
        ])
        ex.export_xlsx(self.path, _sample_invoices(),
                       {'includeRemark': False, 'columns': cols})
        wb = load_workbook(self.path)
        ws = wb['发票汇总']
        headers = [c.value for c in ws[1]]
        self.assertEqual(headers, ['发票号码', '购买方名称', '金额', '税前金额'])
        # 第 1 数据行发票号（合并区左上角，保留值）
        self.assertEqual(ws.cell(row=2, column=1).value, 'INV-001')
        # 第 2 行是 INV-001 的第 2 明细：发票号列已合并（非左上角为空），
        # 改用非合并的「金额」列验证 lineAmount=200
        self.assertEqual(ws.cell(row=3, column=3).value, 200)
        # 第 3 数据行 = INV-002（单明细组，无合并）
        self.assertEqual(ws.cell(row=4, column=1).value, 'INV-002')

    def test_totals_invoice_dedup_and_line_sum(self):
        cols = ex.sanitize_columns([
            {'key': 'invoiceNumber', 'label': '发票号码'},
            {'key': 'lineAmount', 'label': '金额'},       # line 级
            {'key': 'amountWithoutTax', 'label': '税前金额'}, # invoice 级
        ])
        ex.export_xlsx(self.path, _sample_invoices(),
                       {'includeRemark': False, 'columns': cols})
        wb = load_workbook(self.path)
        ws = wb['发票汇总']
        # 合计行 = 第 5 行（3 数据 + 1 合计）
        total_row = [c.value for c in ws[5]]
        # 金额(line) 应为 100+200+50 = 350
        self.assertEqual(total_row[1], 350.0)
        # 税前金额(invoice) 应为 300+50 = 350
        self.assertEqual(total_row[2], 350.0)
        self.assertEqual(ws.cell(row=5, column=1).value, '合计')

    def test_backward_compat_no_columns(self):
        ex.export_xlsx(self.path, _sample_invoices(), {'includeRemark': True})
        wb = load_workbook(self.path)
        ws = wb['发票汇总']
        headers = [c.value for c in ws[1]]
        self.assertIn('发票号码', headers)
        self.assertIn('备注', headers)
        self.assertIn('原文件名', headers)


if __name__ == '__main__':
    unittest.main(verbosity=2)
